#!/usr/bin/env python3
"""
splice_report_generator_without_uni.py — Splice QC report (BIDIRECTIONAL ONLY)
================================================================================

Same as splice_report_generator.py but EXCLUDES all unidirectional entries.
Only fibers with confirmed bidirectional measurements are flagged.

  - Rows: one per 12-fiber ribbon
  - Columns: one per splice closure position
  - Cells: fiber number + bidirectional loss for any fiber requiring attention
  - Breaks: identified by 1F reflective event type (Fresnel = glass-to-air)
  - Reburns: any splice with BIDIRECTIONAL loss >= threshold (default 0.150 dB)
  - Broke: fibers whose trace terminates before the end of span
  - NO unidirectional entries

USAGE
-----
    python splice_report_generator_without_uni.py A_DIR/ B_DIR/ --output report.xlsx

OPTIONS
    --output PATH    Output Excel file (default: splice_report.xlsx)
    --threshold dB   Reburn flag threshold (default 0.150)
    --nominal dB     Nominal splice loss (default 0.159)
    --fibers N       Total fiber count (default: auto-detect)
    --ribbon-size N  Fibers per ribbon (default: 12)

REQUIREMENTS
    pip install numpy openpyxl
    sor_reader324802a.py must be in same directory or on PYTHONPATH.
"""

import os
import sys
import argparse
from collections import defaultdict

import numpy as np

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)

from sor_reader324802a import parse_sor_full


# ═══════════════════════════════════════════════════════════════════════
#  DEFAULTS
# ═══════════════════════════════════════════════════════════════════════

REBURN_THRESHOLD = 0.150   # dB bidirectional — flag anything at or above
NOMINAL_SPLICE   = 0.159   # dB expected per splice
RIBBON_SIZE      = 12      # fibers per ribbon
POSITION_TOL     = 1.5     # km tolerance for matching A↔B events
MIN_POP_SPLICE   = 20      # minimum fibers to define a splice position
END_REGION_KM    = 3.0     # last N km considered "end of fiber"


# ═══════════════════════════════════════════════════════════════════════
#  STEP 1 — Load all fibers
# ═══════════════════════════════════════════════════════════════════════

def load_all(dir_a, dir_b):
    fibers_a, fibers_b = {}, {}
    def extract_fiber_num(fn):
        """Extract fiber number from filenames like STRROM0001_1550.sor or ROMSTR001_1550.sor"""
        base = fn.split('.')[0]           # drop extension
        base = base.split('_')[0]         # drop _1550 wavelength suffix
        digits = ''.join(c for c in base if c.isdigit())
        return int(digits) if digits else None

    for fn in sorted(os.listdir(dir_a)):
        if not fn.lower().endswith('.sor'): continue
        r = parse_sor_full(os.path.join(dir_a, fn))
        if r:
            fnum = extract_fiber_num(fn)
            if fnum: fibers_a[fnum] = r
    if dir_b and os.path.isdir(dir_b):
        for fn in sorted(os.listdir(dir_b)):
            if not fn.lower().endswith('.sor'): continue
            r = parse_sor_full(os.path.join(dir_b, fn))
            if r:
                fnum = extract_fiber_num(fn)
                if fnum: fibers_b[fnum] = r
    return fibers_a, fibers_b


# ═══════════════════════════════════════════════════════════════════════
#  STEP 2 — Discover splice closure positions from the population
# ═══════════════════════════════════════════════════════════════════════

def discover_splices(fibers_a):
    bins = defaultdict(list)
    for fnum, r in fibers_a.items():
        for e in r['events']:
            if e['dist_km'] < 1.0 or e['is_end']: continue
            if not e['type'].startswith('0F') and not e['type'].startswith('1F'): continue
            bk = round(e['dist_km'])
            bins[bk].append(e['dist_km'])

    splices = []
    for bk in sorted(bins.keys()):
        if len(bins[bk]) < MIN_POP_SPLICE: continue
        avg_pos = round(np.mean(bins[bk]), 2)
        splices.append({'bin': bk, 'position_km': avg_pos, 'count': len(bins[bk])})

    # Merge bins that are within 1 km of each other (dedup)
    merged = []
    for sp in splices:
        if merged and abs(sp['position_km'] - merged[-1]['position_km']) < 1.0:
            # Keep the one with more fibers
            if sp['count'] > merged[-1]['count']:
                merged[-1] = sp
        else:
            merged.append(sp)

    return merged


# ═══════════════════════════════════════════════════════════════════════
#  STEP 3 — Analyze every fiber at every splice
# ═══════════════════════════════════════════════════════════════════════

def analyze_all(fibers_a, fibers_b, splices, threshold):
    """
    For each fiber at each splice position:
      - Find the A-direction event nearest to the splice position
      - Find the matching B-direction event (reversed)
      - Compute bidirectional loss = (A + B) / 2
      - Check if event type is 1F (break)
      - Check if fiber terminates before this splice (broke)

    Returns: dict of (fiber, splice_idx) -> result dict
    """
    results = {}  # (fnum, splice_idx) -> {...}

    # Get end-of-fiber distances for broke detection
    eof_a = {}
    for fnum, r in fibers_a.items():
        end = [e for e in r['events'] if e['is_end']]
        eof_a[fnum] = end[0]['dist_km'] if end else 999

    # Auto-detect span length: use the top 25% of EOL distances (full-span fibers)
    # This works regardless of route length — no hardcoded minimum.
    eof_a_vals = sorted(eof_a.values())
    if eof_a_vals:
        top_quarter_a = eof_a_vals[int(len(eof_a_vals) * 0.75):]
        total_span_a = np.median(top_quarter_a)
    else:
        total_span_a = 0

    eof_b = {}
    for fnum, r in fibers_b.items():
        end = [e for e in r['events'] if e['is_end']]
        eof_b[fnum] = end[0]['dist_km'] if end else 999

    eof_b_vals = sorted([v for v in eof_b.values() if v < 999])
    if eof_b_vals:
        top_quarter_b = eof_b_vals[int(len(eof_b_vals) * 0.75):]
        total_span_b = np.median(top_quarter_b)
    else:
        total_span_b = 0

    for fnum, r in fibers_a.items():
        # Get B data
        rb = fibers_b.get(fnum)
        b_span = None
        if rb:
            b_end = [e for e in rb['events'] if e['is_end']]
            b_span = b_end[0]['dist_km'] if b_end else total_span_b

        for si, sp in enumerate(splices):
            sp_km = sp['position_km']

            # Check if fiber is broke BEFORE this splice
            fiber_end = eof_a[fnum]
            # Also check B direction
            fiber_end_b_from_a = (b_span - eof_b.get(fnum, 999)) if fnum in eof_b and b_span else None

            # A fiber is "broke" if its trace terminates early AND the sum
            # of A_end + B_end ≈ total_span (not 2× total_span).
            # Normal fibers: A_end ≈ span, B_end ≈ span → sum ≈ 2×span
            # Broke fibers:  A_end + B_end ≈ 1×span (each side sees to the break)
            a_plus_b = fiber_end + eof_b.get(fnum, 0) if fnum in eof_b else 0
            is_mid_span_break = (a_plus_b > 0 and
                                 abs(a_plus_b - total_span_a) < 3.0 and
                                 fiber_end < total_span_a - END_REGION_KM)

            if is_mid_span_break:
                # Fiber terminated mid-span — mark as broke at the
                # splice nearest to where it terminated (within 2 km)
                nearest_splice = min(range(len(splices)),
                                     key=lambda i: abs(splices[i]['position_km'] - fiber_end))
                nearest_dist = abs(splices[nearest_splice]['position_km'] - fiber_end)
                if nearest_splice == si and nearest_dist < 2.0:
                    results[(fnum, si)] = {
                        'fiber': fnum,
                        'splice_idx': si,
                        'bidir_loss': None,
                        'a_loss': None,
                        'b_loss': None,
                        'bidir_dist': fiber_end,
                        'is_break': False,
                        'is_broke': True,
                        'is_bfill': False,
                        'is_flagged': True,
                        'event_type': 'BROKE',
                        'label': f"{fnum} broke",
                    }
                # For splices PAST the break, use B-direction fill
                elif sp_km > fiber_end and rb and b_span:
                    # Find B event near this splice position
                    b_evt = None
                    for e in rb['events']:
                        if e['dist_km'] < 1.0 or e['is_end']:
                            continue
                        ef_from_a = b_span - e['dist_km']
                        if abs(ef_from_a - sp_km) < POSITION_TOL:
                            if b_evt is None or abs(ef_from_a - sp_km) < abs((b_span - b_evt['dist_km']) - sp_km):
                                b_evt = e
                    if b_evt is not None:
                        b_loss_val = abs(b_evt['splice_loss'])
                        if b_loss_val >= threshold:
                            loss_str = f"{b_loss_val:.3f}"
                            if loss_str.startswith('0.'):
                                loss_str = loss_str[1:]
                            results[(fnum, si)] = {
                                'fiber': fnum,
                                'splice_idx': si,
                                'bidir_loss': b_loss_val,
                                'a_loss': None,
                                'b_loss': b_evt['splice_loss'],
                                'bidir_dist': b_span - b_evt['dist_km'],
                                'is_break': False,
                                'is_broke': False,
                                'is_bfill': True,
                                'is_flagged': True,
                                'event_type': b_evt['type'],
                                'label': f"{fnum} {loss_str} (B)",
                            }
                continue

            # Find A event near this splice
            ea = None
            for e in r['events']:
                if abs(e['dist_km'] - sp_km) < POSITION_TOL and e['dist_km'] > 1.0 and not e['is_end']:
                    if ea is None or abs(e['dist_km'] - sp_km) < abs(ea['dist_km'] - sp_km):
                        ea = e

            if ea is None:
                continue

            # Find matching B event
            eb = None
            b_loss = None
            b_from_a = None
            if rb and b_span:
                for e in rb['events']:
                    if e['dist_km'] < 1.0 or e['is_end']: continue
                    ef_from_a = b_span - e['dist_km']
                    if abs(ef_from_a - ea['dist_km']) < POSITION_TOL:
                        if eb is None or abs(ef_from_a - ea['dist_km']) < abs((b_span - eb['dist_km']) - ea['dist_km']):
                            eb = e
                            b_loss = e['splice_loss']
                            b_from_a = ef_from_a

            # Bidirectional calculations — SKIP unidirectional entirely
            is_unidirectional = (b_loss is None)
            if is_unidirectional:
                continue  # ← without_uni: exclude all unidirectional entries

            bidir_loss = round((ea['splice_loss'] + b_loss) / 2.0, 4)
            bidir_dist = round((ea['dist_km'] + b_from_a) / 2.0, 4)

            is_reflective = ea['type'].startswith('1F')
            has_weak_fresnel = ea['reflection'] < -30.0
            is_break = is_reflective and has_weak_fresnel and ea['dist_km'] < (total_span_a - END_REGION_KM)

            is_flagged = (abs(bidir_loss) >= threshold) or is_break

            if not is_flagged:
                continue

            # Build label
            if is_break:
                offset_m = round((bidir_dist - sp_km) * 1000, 1)
                offset_ft = round(offset_m * 3.28084, 0)
                label = f"{fnum} BREAK {bidir_loss:.3f} ({abs(offset_m):.0f}m from splice)"
            else:
                loss_str = f"{bidir_loss:.3f}"
                if loss_str.startswith('0.'):
                    loss_str = loss_str[1:]
                label = f"{fnum} {loss_str}"

            results[(fnum, si)] = {
                'fiber': fnum,
                'splice_idx': si,
                'bidir_loss': bidir_loss,
                'a_loss': ea['splice_loss'],
                'b_loss': b_loss,
                'bidir_dist': bidir_dist,
                'is_break': is_break,
                'is_broke': False,
                'is_bfill': False,
                'is_flagged': True,
                'event_type': ea['type'],
                'label': label,
                'fresnel': ea['reflection'] if is_reflective else None,
            }

    return results


# ═══════════════════════════════════════════════════════════════════════
#  STEP 4 — Group into ribbons and build cell values
# ═══════════════════════════════════════════════════════════════════════

def build_ribbon_data(results, n_fibers, ribbon_size, n_splices):
    """Group flagged events by ribbon and splice position."""
    n_ribbons = (n_fibers + ribbon_size - 1) // ribbon_size
    grid = {}  # (ribbon_idx, splice_idx) -> cell text

    for (fnum, si), res in results.items():
        ri = (fnum - 1) // ribbon_size
        key = (ri, si)
        if key not in grid:
            grid[key] = []
        grid[key].append(res)

    # Merge multiple fibers in same cell
    cells = {}
    for (ri, si), res_list in grid.items():
        # Group fibers with same loss (ribbons often have multiple fibers at same loss)
        # Sort by fiber number
        res_list.sort(key=lambda x: x['fiber'])

        # Check if multiple fibers share the same loss (within 0.001)
        groups = []
        for res in res_list:
            merged = False
            for g in groups:
                if (res['bidir_loss'] is not None and g['loss'] is not None and
                        abs(res['bidir_loss'] - g['loss']) < 0.002 and
                        not res['is_break'] and not res['is_broke'] and
                        not g['is_break'] and not g['is_broke']):
                    g['fibers'].append(res['fiber'])
                    merged = True
                    break
            if not merged:
                groups.append({
                    'fibers': [res['fiber']],
                    'loss': res['bidir_loss'],
                    'is_break': res['is_break'],
                    'is_broke': res['is_broke'],
                    'is_bfill': res.get('is_bfill', False),
                    'label': res['label'],
                    'res': res,
                })

        # Build cell text
        parts = []
        for g in groups:
            if g['is_broke']:
                parts.append(f"{g['fibers'][0]} broke")
            elif g['is_break']:
                parts.append(g['label'])
            else:
                fib_str = ','.join(str(f) for f in g['fibers'])
                loss = g['loss']
                loss_str = f"{loss:.3f}" if loss is not None else "?"
                if loss_str.startswith('0.'):
                    loss_str = loss_str[1:]
                parts.append(f"{fib_str} {loss_str}")

        cell_text = ' '.join(parts)
        is_break = any(g['is_break'] for g in groups)
        is_broke = any(g['is_broke'] for g in groups)
        is_bfill = any(g.get('is_bfill', False) for g in groups)
        max_loss = max((g['loss'] for g in groups if g['loss'] is not None), default=0)

        cells[(ri, si)] = {
            'text': cell_text,
            'is_break': is_break,
            'is_broke': is_broke,
            'is_bfill': is_bfill,
            'max_loss': max_loss,
        }

    return cells


# ═══════════════════════════════════════════════════════════════════════
#  STEP 5 — Generate Excel
# ═══════════════════════════════════════════════════════════════════════

def ribbon_label(ri, ribbon_size, n_fibers):
    """Generate ribbon label like 'Fiber 1-12 (1) (A1)'"""
    first = ri * ribbon_size + 1
    last = min(first + ribbon_size - 1, n_fibers)
    ribbon_num = ri + 1

    # Tube naming: A through X for first 48 ribbons (24 tubes × 2 ribbons)
    tube = ''
    if ri < 48:
        tube_letter = chr(ord('A') + ri // 2)
        tube_num = (ri % 2) + 1
        tube = f" ({tube_letter}{tube_num})"

    return f"Fiber {first}-{last} ({ribbon_num}){tube}"


def write_xlsx(cells, splices, n_fibers, ribbon_size, output_path, site_a, site_b, span_km):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Splice Report"

    n_ribbons = (n_fibers + ribbon_size - 1) // ribbon_size
    n_splices = len(splices)

    # ── Styles ──
    hdr_font = Font(bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    dist_font = Font(bold=True, size=9, color="1F4E79")
    ribbon_font = Font(size=9)
    data_font = Font(size=8)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    break_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
    break_font = Font(bold=True, size=8, color="FFFFFF")
    broke_fill = PatternFill(start_color="FF8800", end_color="FF8800", fill_type="solid")
    broke_font = Font(bold=True, size=8, color="FFFFFF")
    bfill_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    bfill_font = Font(size=8, color="1F4E79")
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # ── Row 1: A→B distances (km / ft) ──
    a_km_font = Font(bold=True, size=9, color="1F4E79")
    b_km_font = Font(bold=True, size=9, color="8B0000")
    ws.cell(row=1, column=2, value="A→B:").font = a_km_font
    ws.cell(row=2, column=2, value="B→A:").font = b_km_font
    for si, sp in enumerate(splices):
        col = si + 3  # columns C onward
        km = sp['position_km']
        ft = km * 3280.84
        b_km = span_km - km
        b_ft = b_km * 3280.84
        c1 = ws.cell(row=1, column=col, value=f"{km:.2f}km / {ft:,.0f}ft")
        c1.font = a_km_font
        c1.alignment = Alignment(horizontal='center')
        c2 = ws.cell(row=2, column=col, value=f"{b_km:.2f}km / {b_ft:,.0f}ft")
        c2.font = b_km_font
        c2.alignment = Alignment(horizontal='center')
    # Last column: end site
    end_col = n_splices + 3
    c1 = ws.cell(row=1, column=end_col, value=f"{span_km:.2f}km / {span_km*3280.84:,.0f}ft")
    c1.font = a_km_font
    c2 = ws.cell(row=2, column=end_col, value="0.00km / 0ft")
    c2.font = b_km_font

    # ── Row 3: Headers ──
    ws.cell(row=3, column=1, value="Ribbon").font = hdr_font
    ws.cell(row=3, column=1).fill = hdr_fill
    ws.cell(row=3, column=2, value=f"ILA:{site_a}").font = hdr_font
    ws.cell(row=3, column=2).fill = hdr_fill
    for si in range(n_splices):
        col = si + 3
        cell = ws.cell(row=3, column=col, value=f"Splice {si+1}")
        cell.font = hdr_font
        cell.fill = hdr_fill
    cell = ws.cell(row=3, column=end_col, value=f"ILA: {site_b}")
    cell.font = hdr_font
    cell.fill = hdr_fill

    # ── Data rows ──
    for ri in range(n_ribbons):
        row = ri + 4
        # Column A: ribbon label
        ws.cell(row=row, column=1, value=ribbon_label(ri, ribbon_size, n_fibers)).font = ribbon_font

        # Splice columns
        for si in range(n_splices):
            col = si + 3
            key = (ri, si)
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')

            if key in cells:
                cd = cells[key]
                cell.value = cd['text']
                cell.font = data_font

                if cd['is_break']:
                    cell.fill = break_fill
                    cell.font = break_font
                elif cd['is_broke']:
                    cell.fill = broke_fill
                    cell.font = broke_font
                elif cd.get('is_bfill', False):
                    cell.fill = bfill_fill
                    cell.font = bfill_font
                else:
                    cell.fill = red_fill
                    cell.font = data_font

    # ── Column widths ──
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 10
    for si in range(n_splices + 1):
        col_letter = openpyxl.utils.get_column_letter(si + 3)
        ws.column_dimensions[col_letter].width = 22

    # ── Freeze panes ──
    ws.freeze_panes = 'C3'

    wb.save(output_path)
    print(f"  Saved: {output_path}")


# ═══════════════════════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description='Generate splice QC report from raw OTDR SOR data.',
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('dir_a', help='A-direction SOR files directory')
    ap.add_argument('dir_b', nargs='?', help='B-direction SOR files directory (optional)')
    ap.add_argument('--output', '-o', default='splice_report.xlsx', help='Output Excel file')
    ap.add_argument('--threshold', type=float, default=REBURN_THRESHOLD,
                    help=f'Reburn flag threshold in dB (default {REBURN_THRESHOLD})')
    ap.add_argument('--nominal', type=float, default=NOMINAL_SPLICE,
                    help=f'Nominal splice loss (default {NOMINAL_SPLICE})')
    ap.add_argument('--ribbon-size', type=int, default=RIBBON_SIZE,
                    help=f'Fibers per ribbon (default {RIBBON_SIZE})')
    ap.add_argument('--site-a', default='STR', help='Site A name (default STR)')
    ap.add_argument('--site-b', default='ROM', help='Site B name (default ROM)')
    ap.add_argument('--span-km', type=float, default=0, help='Span distance in km (auto-detect if 0)')
    args = ap.parse_args()

    print(f"Loading SOR files...")
    fibers_a, fibers_b = load_all(args.dir_a, args.dir_b)
    n_fibers = max(fibers_a.keys()) if fibers_a else 0
    print(f"  A: {len(fibers_a)} fibers, B: {len(fibers_b)} fibers, max fiber #{n_fibers}")

    print(f"Discovering splice positions...")
    splices = discover_splices(fibers_a)
    print(f"  Found {len(splices)} splice closures:")
    for i, sp in enumerate(splices, 1):
        print(f"    Splice {i:2d}: {sp['position_km']:8.2f} km ({sp['count']} fibers)")

    # Auto-detect span
    span_km = args.span_km
    if span_km == 0:
        all_ends = sorted([e['dist_km'] for r in fibers_a.values()
                           for e in r['events'] if e['is_end']])
        if all_ends:
            top_quarter = all_ends[int(len(all_ends) * 0.75):]
            span_km = round(np.median(top_quarter), 2)
        else:
            span_km = 0
    print(f"  Span: {span_km} km")

    print(f"Analyzing {len(fibers_a)} fibers at {len(splices)} splices (threshold={args.threshold:.3f} dB)...")
    results = analyze_all(fibers_a, fibers_b, splices, args.threshold)
    n_flagged = len(results)
    n_breaks = sum(1 for r in results.values() if r['is_break'])
    n_broke = sum(1 for r in results.values() if r['is_broke'])
    n_bfill = sum(1 for r in results.values() if r.get('is_bfill', False))
    n_reburn = n_flagged - n_breaks - n_broke - n_bfill
    print(f"  Flagged: {n_flagged} events ({n_breaks} breaks, {n_broke} broke, {n_bfill} B-fill, {n_reburn} reburns)")

    print(f"Building ribbon grid...")
    cells = build_ribbon_data(results, n_fibers, args.ribbon_size, len(splices))
    print(f"  {len(cells)} cells with data")

    print(f"Writing Excel report...")
    write_xlsx(cells, splices, n_fibers, args.ribbon_size, args.output,
               args.site_a, args.site_b, span_km)

    # Summary
    print(f"\n{'═'*60}")
    print(f"  SPLICE REPORT GENERATED")
    print(f"{'═'*60}")
    print(f"  Fibers:     {n_fibers}")
    print(f"  Splices:    {len(splices)}")
    print(f"  Threshold:  {args.threshold:.3f} dB")
    print(f"  Breaks:     {n_breaks}")
    print(f"  Broke:      {n_broke}")
    print(f"  B-fill:     {n_bfill}")
    print(f"  Reburns:    {n_reburn}")
    print(f"  Output:     {args.output}")
    print()


if __name__ == '__main__':
    main()
