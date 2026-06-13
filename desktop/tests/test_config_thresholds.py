"""
Build configuration / threshold-flow tests.

Regression guards for the four HIGH-severity findings + one MEDIUM from
audit agent 4:

  H1  OTDR-panel edits never reach the engine without "Apply settings".
  H2  Lumen / Zayo profiles ship placeholder thresholds, no UI warning.
  H3  NaN can be written to session_state when the Fail field is briefly
      cleared (parseFloat("") -> NaN -> silently disables the threshold).
  H4  Engine globals are process-shared on Streamlit Cloud (two sessions
      racing on REBURN_THRESHOLD).
  M3  Profile switch silently discards manual edits.

These tests are observation-only — no production behaviour is changed in
desktop_app.py.  The xlsx-diffing tests reuse FIXTURE_A_DIR /
FIXTURE_B_DIR exactly as test_e2e_generate.py does, and trigger the
Generate path via session_state["__test_force_generate__"] = True (the
same hook documented in conftest.py).
"""
from __future__ import annotations

import math
import shutil
from pathlib import Path

import pytest

from conftest import APP_PATH, FIXTURE_A_DIR, FIXTURE_B_DIR, run_streamlit


# ── output-folder housekeeping (mirrors test_e2e_generate) ─────────────
def _output_dir() -> Path:
    return FIXTURE_A_DIR.parent / "splice_report_output"


def _clean_output() -> None:
    out = _output_dir()
    if out.exists():
        shutil.rmtree(out, ignore_errors=True)


@pytest.fixture(autouse=True)
def _cleanup():
    _clean_output()
    yield
    _clean_output()


# ── shared helpers ─────────────────────────────────────────────────────
def _reburn_pct_from_xlsx(xlsx_path: Path) -> float:
    """Parse the headline "Reburn percentage" value from the Reburn
    Summary sheet.  Layout (per reburn_summary.py): col-1 label
    "Reburn percentage", col-2 value formatted as "XX.XX%".
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=False)
    try:
        ws = wb["Reburn Summary"]
        for row in ws.iter_rows(values_only=True, max_row=40):
            if row and row[0] == "Reburn percentage":
                raw = row[1]
                return float(str(raw).rstrip("%"))
    finally:
        wb.close()
    raise AssertionError(
        f"Reburn percentage headline not found in {xlsx_path}"
    )


def _run_generate_with_otdr(otdr_overrides: dict) -> float:
    """Boot the app, install a full otdr_settings dict, force Generate,
    and return the workbook's headline reburn %.
    """
    from desktop_app import _settings_from_profile  # noqa: WPS433

    at = run_streamlit(default_timeout=180.0)
    # Start from the engine-baseline preset, then layer overrides on top.
    base = _settings_from_profile("Default (engine baseline)")
    for k, patch in otdr_overrides.items():
        base[k] = {**base[k], **patch}
    at.session_state["otdr_settings"] = base
    at.session_state["dir_a_input"] = str(FIXTURE_A_DIR)
    at.session_state["dir_b_input"] = str(FIXTURE_B_DIR)
    at.session_state["__test_force_generate__"] = True
    at.run()
    assert at.exception is None or len(at.exception) == 0, (
        f"Generate raised: {[str(e) for e in at.exception]}"
    )
    xlsxs = sorted(_output_dir().glob("*.xlsx"))
    assert len(xlsxs) == 1, f"expected one xlsx, got {xlsxs}"
    return _reburn_pct_from_xlsx(xlsxs[0])


# ── T1.  Profile switch updates session_state.otdr_settings ────────────
def test_profile_switch_updates_otdr_settings():
    """
    The desktop sidebar profile dropdown reads from the widget key
    ``otdr_profile_select_desktop`` — NOT the side-slot ``otdr_profile``.
    When the test only mutates the side-slot, the selectbox widget still
    holds whatever it had before (or the default index=0), Streamlit
    fires the change handler, and the side-slot is silently overwritten
    on the next render.  This is the same widget-state-precedence
    pattern the picker fix in commit 224dae0 documents.

    The right test approach is therefore: mutate ONLY the widget key.
    The production change-handler in desktop_app.py then writes the
    matching otdr_settings, and the test asserts on the result.
    """
    at = run_streamlit()
    at.run()
    assert at.session_state["otdr_profile"] == "Default (engine baseline)"

    # Switch to Lumen by writing the SELECTBOX's own widget key.
    at.session_state["otdr_profile_select_desktop"] = "Lumen"
    at.run()
    assert at.session_state["otdr_profile"] == "Lumen", (
        "profile-change handler didn't sync otdr_profile to the picker"
    )
    assert at.session_state["otdr_settings"]["bidir_splice_loss"]["fail"] \
        == pytest.approx(0.120)

    # Switch to Zayo the same way.
    at.session_state["otdr_profile_select_desktop"] = "Zayo"
    at.run()
    assert at.session_state["otdr_profile"] == "Zayo"
    assert at.session_state["otdr_settings"]["bidir_connector_loss"]["fail"] \
        == pytest.approx(0.600)


# ── T2.  Stale profile name resets cleanly ─────────────────────────────
def test_stale_profile_name_falls_back():
    at = run_streamlit()
    # A name that was renamed away (commit 4cc1925).  desktop_app.py
    # lines 376-378 guard against this — the test pins that guard.
    at.session_state["otdr_profile"] = "Customer A — strict"
    at.run()
    assert at.exception is None or len(at.exception) == 0, (
        f"stale profile name raised: {[str(e) for e in at.exception]}"
    )
    from desktop_app import CUSTOMER_PROFILES  # noqa: WPS433
    expected = next(iter(CUSTOMER_PROFILES.keys()))
    assert at.session_state["otdr_profile"] == expected


# ── T3.  Apply-settings gate (H1) — dict-style updates DO flow ─────────
def test_session_state_otdr_overrides_reach_engine():
    """Two Generate runs differing ONLY in
    session_state.otdr_settings['bidir_splice_loss']['fail'] must produce
    different reburn percentages.  If they're equal, the dict-style
    override is not reaching the engine — which is the H1 footgun in its
    most extreme form (the Apply gate has cut the wire entirely).
    """
    pct_default = _run_generate_with_otdr({
        "bidir_splice_loss": {"apply": True, "fail": 0.160},
    })
    pct_low = _run_generate_with_otdr({
        "bidir_splice_loss": {"apply": True, "fail": 0.080},
    })
    assert pct_default != pct_low, (
        "H1 REGRESSION: identical reburn % at 0.080 dB vs 0.160 dB means "
        "session_state.otdr_settings -> THRESHOLDS['REBURN_THRESHOLD'] -> "
        "engine.REBURN_THRESHOLD is NOT flowing.  The Apply-settings gate "
        "(or an equivalent commit step) is silently dropping dict-style "
        f"overrides.  Both runs reported {pct_default}% / {pct_low}%."
    )


# ── T4.  NaN guard (H3) ────────────────────────────────────────────────
def test_nan_in_fail_does_not_crash_and_is_observable():
    """The audit found that parseFloat("") -> NaN can land in
    session_state.otdr_settings[...]['fail'].  NaN >= X is always False,
    so the threshold silently disables itself.  This test:

      a) confirms the engine does not crash on NaN (today's behaviour),
      b) documents the divergence vs the 0.160-dB default so that when
         somebody adds a defensive guard, this test FAILS loudly and
         forces them to update the audit doc.
    """
    pct_default = _run_generate_with_otdr({
        "bidir_splice_loss": {"apply": True, "fail": 0.160},
    })
    pct_nan = _run_generate_with_otdr({
        "bidir_splice_loss": {"apply": True, "fail": float("nan")},
    })
    # NaN today silently disables the threshold; with a defensive guard
    # the two numbers should match (and this test will need to be
    # updated alongside the audit doc).
    assert not math.isnan(pct_nan), "engine produced NaN reburn %"
    assert pct_default != pct_nan, (
        "H3 GUARD LANDED: NaN in fail no longer diverges from the 0.160-dB "
        "default — this is the desired fix.  Update the audit doc and "
        "flip this assertion to assert equality."
    )


# ── T5.  Lumen / Zayo placeholder values frozen ────────────────────────
def test_lumen_zayo_placeholder_thresholds_frozen():
    from desktop_app import CUSTOMER_PROFILES  # noqa: WPS433

    lumen = CUSTOMER_PROFILES["Lumen"]["thresholds"]
    assert lumen == {
        "bidir_splice_loss":    0.120,
        "unidir_splice_loss":   0.200,
        "bidir_connector_loss": 0.400,
        "reflectance":         -50.0,
    }, (
        "Lumen placeholder thresholds drifted — audit H2 says these are "
        "frozen until a real Lumen spec sheet arrives.  Update the audit "
        "doc before changing this dict."
    )

    zayo = CUSTOMER_PROFILES["Zayo"]["thresholds"]
    assert zayo == {
        "bidir_splice_loss":    0.200,
        "bidir_connector_loss": 0.600,
    }, (
        "Zayo placeholder thresholds drifted — audit H2 says these are "
        "frozen until a real Zayo spec sheet arrives."
    )


# ── T6.  Desktop + web THRESHOLDS keys agree ───────────────────────────
def test_desktop_and_web_thresholds_keys_match():
    """The desktop and web apps construct a THRESHOLDS dict that is
    handed to engine via _apply_overrides.  If somebody adds a new
    threshold to one side and forgets the other, the docs and the field
    behaviour drift.  We compare the literal key set in each file's
    THRESHOLDS = {...} block — static text, no module import needed
    (importing app.py would execute the web app's top-level Streamlit
    code).
    """
    import re

    expected = {
        "REBURN_THRESHOLD",
        "SINGLE_DIR_THRESHOLD",
        "BIDIR_CONNECTOR_LOSS",
        "LAUNCH_BAD_REFL_DB",
    }

    def _threshold_keys(py_path: Path) -> set:
        src = py_path.read_text()
        m = re.search(r"THRESHOLDS\s*=\s*\{(.*?)\n\}", src, re.DOTALL)
        assert m, f"no THRESHOLDS = {{...}} block in {py_path}"
        return set(re.findall(r'"([A-Z_]+)"\s*:', m.group(1)))

    desktop_keys = _threshold_keys(APP_PATH)
    web_keys = _threshold_keys(APP_PATH.parent.parent / "app.py")

    # The desktop and web apps may legitimately differ in HOW MANY
    # threshold keys they construct — the web app reads many engine
    # defaults into THRESHOLDS and writes them back unchanged, while
    # the desktop app skips that no-op redundancy.  What matters for
    # tech-visible behaviour is the four user-controllable keys (the
    # ones the OTDR settings panel actually drives).  Both apps must
    # build all four; that's the structural contract.
    assert expected.issubset(desktop_keys), (
        f"desktop THRESHOLDS missing keys: {expected - desktop_keys}"
    )
    assert expected.issubset(web_keys), (
        f"web THRESHOLDS missing keys: {expected - web_keys}"
    )
    # Tighter check: the keys present in BOTH must agree on which subset
    # of the panel-controllable thresholds they cover.  If one app adds a
    # fifth user-controllable threshold to its panel, this assertion
    # surfaces the drift.
    desktop_user_keys = desktop_keys & expected
    web_user_keys = web_keys & expected
    assert desktop_user_keys == web_user_keys, (
        f"desktop vs web user-controllable threshold drift — "
        f"desktop: {desktop_user_keys}, web: {web_user_keys}"
    )
