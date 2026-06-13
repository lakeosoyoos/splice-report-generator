"""
End-to-end "Generate report" test.

This is the gap-closer for the class of bug that shipped in commit
e1d5692: ``ValueError: too many values to unpack (expected 2)`` at the
``detect_launch_issues`` call site in desktop_app.py.  None of our
existing CI checks caught it — the boot self-test only confirmed
``/_stcore/health=ok`` (the server was up before the user clicked
Generate), and the AppTest smoke tests in desktop/test_ui.py exercised
widget-state plumbing but never actually ran the engine pipeline.

This test:
    1. Points an AppTest at desktop_app.py.
    2. Feeds it the bundled tiny fixture span (desktop/tests/fixtures/).
    3. Sets ``session_state["__test_force_generate__"] = True`` to
       trigger the Generate-report code path without simulating a click.
    4. Asserts the run finishes cleanly (``at.exception is None``).
    5. Asserts a workbook lands in splice_report_output/ next to the A
       input, with the expected four sheets, the expected active sheet,
       and populated Reburn Summary / Splice Report header cells.

If the UI and the engine drift apart in signature (the e1d5692-shape
bug), step 4 fails with the underlying ValueError.
"""
from __future__ import annotations

import shutil
from pathlib import Path

import pytest

from conftest import (
    APP_PATH,
    FIXTURE_A_DIR,
    FIXTURE_B_DIR,
    run_streamlit,
)


def _output_dir() -> Path:
    """Where desktop_app.py writes its xlsx.

    The UI computes::

        out_dir = os.path.join(os.path.dirname(os.path.abspath(dir_a)),
                               "splice_report_output")

    For our fixture that's ``desktop/tests/fixtures/splice_report_output``.
    """
    return FIXTURE_A_DIR.parent / "splice_report_output"


def _clean_output() -> None:
    out = _output_dir()
    if out.exists():
        shutil.rmtree(out, ignore_errors=True)


@pytest.fixture(autouse=True)
def _cleanup():
    """Wipe the output folder before and after each test."""
    _clean_output()
    yield
    _clean_output()


def test_fixtures_exist():
    """Sanity check — the bundled fixture span is present."""
    assert APP_PATH.is_file(), f"missing desktop_app.py at {APP_PATH}"
    assert FIXTURE_A_DIR.is_dir(), f"missing A fixture at {FIXTURE_A_DIR}"
    assert FIXTURE_B_DIR.is_dir(), f"missing B fixture at {FIXTURE_B_DIR}"
    a_sors = sorted(FIXTURE_A_DIR.glob("*.sor"))
    b_sors = sorted(FIXTURE_B_DIR.glob("*.sor"))
    # 36 fibers per side = 3 ribbons of 12.  The engine's
    # discover_splices gate requires MIN_POP_SPLICE (=20) fibers in a
    # 1 km bucket before it records a closure, so a single ribbon
    # (12 fibers) is too thin to yield any splices and downstream
    # scan_a_standalone_events blows up indexing an empty list.
    assert len(a_sors) >= 36, f"expected >=36 A-side SOR, got {len(a_sors)}"
    assert len(b_sors) >= 36, f"expected >=36 B-side SOR, got {len(b_sors)}"


def test_generate_report_e2e():
    """The full Generate-report path runs without raising.

    This is the test that would have caught commit e1d5692.
    """
    at = run_streamlit(default_timeout=120.0)

    # Feed the picker text-input slots (the same slot ``test_ui.py``
    # exercises for the dialog-vs-textinput plumbing bug).
    at.session_state["dir_a_input"] = str(FIXTURE_A_DIR)
    at.session_state["dir_b_input"] = str(FIXTURE_B_DIR)

    # Trigger Generate via the test hook (see desktop_app.py).
    at.session_state["__test_force_generate__"] = True

    at.run()

    # ── 1. No exception ─────────────────────────────────────────────
    # This is THE assertion that catches signature drift between
    # desktop_app.py and the engine.  e1d5692 fails here with
    # "too many values to unpack (expected 2)".
    assert at.exception is None or len(at.exception) == 0, (
        f"Generate-report path raised — likely signature drift between "
        f"desktop_app.py and the engine.  Exception(s): "
        f"{[str(e) for e in at.exception]}"
    )

    # ── 2. Workbook written ─────────────────────────────────────────
    out_dir = _output_dir()
    assert out_dir.is_dir(), (
        f"expected workbook output directory {out_dir} to exist"
    )
    xlsx_files = sorted(out_dir.glob("*.xlsx"))
    assert len(xlsx_files) == 1, (
        f"expected exactly one workbook in {out_dir}, found: {xlsx_files}"
    )
    xlsx_path = xlsx_files[0]
    assert xlsx_path.stat().st_size > 0, "workbook is empty"

    # ── 3. Workbook structure ───────────────────────────────────────
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=False, data_only=False)

    expected_order = [
        "Acquisition Parameters",
        "Reburn Summary",
        "Splice Report",
        "Legend",
    ]
    assert wb.sheetnames == expected_order, (
        f"sheet order regressed: got {wb.sheetnames}, "
        f"expected {expected_order}"
    )

    # Active sheet must be Acquisition Parameters (index 0).
    assert wb.active.title == "Acquisition Parameters", (
        f"workbook opens on {wb.active.title!r}, "
        f"expected 'Acquisition Parameters'"
    )

    # ── 4. Reburn Summary headline cells populated ──────────────────
    ws_reburn = wb["Reburn Summary"]
    # Slurp the first column or two looking for the three labels we
    # promise to surface.  Cell layout has shifted before; matching on
    # label text is more durable than hard-coding row/col.
    reburn_text = []
    for row in ws_reburn.iter_rows(values_only=True, max_row=40):
        for cell in row:
            if cell is not None:
                reburn_text.append(str(cell))
    blob = " | ".join(reburn_text).lower()
    for needle in ("ribbon", "total", "reburn"):
        assert needle in blob, (
            f"Reburn Summary missing '{needle}' headline — got: "
            f"{reburn_text[:20]!r}..."
        )

    # ── 5. Splice Report has a header row ───────────────────────────
    ws_splice = wb["Splice Report"]
    first_row = next(
        ws_splice.iter_rows(values_only=True, max_row=1), tuple()
    )
    non_empty = [c for c in first_row if c not in (None, "")]
    assert len(non_empty) > 0, (
        "Splice Report sheet has no header row content"
    )

    wb.close()
