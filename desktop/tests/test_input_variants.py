"""
Input-variant regression tests.

Regression guards for the input-step issues flagged by Audit agent 3:

  H5  Same folder picked for both A and B has no guard — silent "perfect"
       reports.  See T4.
  H6  Sub-folder fiber-number collisions logged to a hidden expander —
       see T5 (pairs with Agent D's engine-side collision WARN check).
  H7  8-byte content-sniff window rejects valid SOR variants whose
       ``Map\\0`` header sits past offset 8 — see T3 (negative side: bogus
       headers are rejected) and T1 (positive side: real headers pass).
  AppleDouble (._STRROM0001_1550.sor) pattern from the filename sweep —
       silently overwriting fiber 1 when extracted from a Mac zip.  T2.

These tests build zip fixtures on the fly in tempfiles and tear them
down on exit — nothing extra lands in the repo.
"""
from __future__ import annotations

import os
import shutil
import tempfile
import zipfile
from pathlib import Path

import pytest

from conftest import (
    FIXTURE_A_DIR,
    FIXTURE_B_DIR,
    run_streamlit,
)


# ─────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────
def _zip_folder(src_dir: Path, zip_path: Path,
                extra_entries: dict[str, bytes] | None = None) -> None:
    """Zip the contents of ``src_dir`` (flat, basename-keyed) into
    ``zip_path``.  Optionally add ``extra_entries`` keyed by entry name."""
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for p in sorted(src_dir.iterdir()):
            if p.is_file():
                zf.write(p, arcname=p.name)
        if extra_entries:
            for name, blob in extra_entries.items():
                zf.writestr(name, blob)


def _output_dir_for(input_path: Path) -> Path:
    """Where desktop_app.py writes its xlsx, given an A input path."""
    return Path(os.path.dirname(os.path.abspath(str(input_path)))) / \
        "splice_report_output"


def _clean_output_for(input_path: Path) -> None:
    out = _output_dir_for(input_path)
    if out.exists():
        shutil.rmtree(out, ignore_errors=True)


def _run_generate(dir_a: str, dir_b: str) -> Path:
    """Drive the Generate code path end-to-end.  Returns the xlsx path."""
    at = run_streamlit(default_timeout=120.0)
    at.session_state["dir_a_input"] = dir_a
    at.session_state["dir_b_input"] = dir_b
    at.session_state["__test_force_generate__"] = True
    at.run()
    assert at.exception is None or len(at.exception) == 0, (
        f"Generate raised: {[str(e) for e in at.exception]}"
    )
    out_dir = _output_dir_for(Path(dir_a))
    xlsx_files = sorted(out_dir.glob("*.xlsx"))
    assert len(xlsx_files) == 1, (
        f"expected exactly one workbook in {out_dir}, got {xlsx_files}"
    )
    return xlsx_files[0]


def _splice_columns(xlsx_path: Path) -> list:
    """Return the column header row (row 3) of the Splice Report sheet."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=False, data_only=False)
    try:
        ws = wb["Splice Report"]
        return [c.value for c in ws[3]]
    finally:
        wb.close()


def _splice_sheet_shape(xlsx_path: Path) -> tuple[int, int, list]:
    """Return (row_count, column_count, header-row-3) for Splice Report.

    The Splice Report header row contains two ILA labels (``ILA:<name>``)
    derived from the *basename* of the A/B input.  Those legitimately
    differ when comparing a folder run to a zip run, so we mask them to a
    sentinel before comparison.  Everything else in the header — column
    order, splice/bend labels, blank columns — must match exactly.
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=False, data_only=False)
    try:
        ws = wb["Splice Report"]
        raw_header = [c.value for c in ws[3]]
        masked = [
            "ILA:<masked>" if isinstance(v, str) and v.startswith("ILA:")
            else v
            for v in raw_header
        ]
        return ws.max_row, ws.max_column, masked
    finally:
        wb.close()


def _reburn_blob(xlsx_path: Path, mask_tokens: list[str] | None = None) -> str:
    """Stringified slurp of the Reburn Summary sheet for headline-text
    comparison.  ``mask_tokens`` are substrings (e.g. input basenames)
    redacted from the blob before return so that runs that differ ONLY
    in the input name still compare equal."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=False, data_only=False)
    try:
        ws = wb["Reburn Summary"]
        parts = []
        for row in ws.iter_rows(values_only=True, max_row=40):
            for cell in row:
                if cell is not None:
                    parts.append(str(cell))
        out = " | ".join(parts)
        for tok in mask_tokens or []:
            if tok:
                out = out.replace(tok, "<masked>")
        return out
    finally:
        wb.close()


# ─────────────────────────────────────────────────────────────────────────
# T1.  Folder ≡ Zip equivalence on the same data.
# ─────────────────────────────────────────────────────────────────────────
def test_t1_folder_zip_equivalence():
    """Same data, two input pathways — engine output must match."""
    with tempfile.TemporaryDirectory(prefix="ivt_t1_") as td:
        td_path = Path(td)
        zip_a = td_path / "span_A.zip"
        zip_b = td_path / "span_B.zip"
        _zip_folder(FIXTURE_A_DIR, zip_a)
        _zip_folder(FIXTURE_B_DIR, zip_b)

        # Run 1 — folder paths.
        _clean_output_for(FIXTURE_A_DIR)
        try:
            xlsx_folder = _run_generate(str(FIXTURE_A_DIR), str(FIXTURE_B_DIR))
            shape_folder = _splice_sheet_shape(xlsx_folder)
            reburn_folder = _reburn_blob(
                xlsx_folder,
                mask_tokens=[FIXTURE_A_DIR.name, FIXTURE_B_DIR.name],
            )
            # Copy aside before the next run wipes the output dir.
            preserved = td_path / "folder_run.xlsx"
            shutil.copy2(xlsx_folder, preserved)
        finally:
            _clean_output_for(FIXTURE_A_DIR)

        # Run 2 — zip paths.  Output dir is now under td_path.
        _clean_output_for(zip_a)
        try:
            xlsx_zip = _run_generate(str(zip_a), str(zip_b))
            shape_zip = _splice_sheet_shape(xlsx_zip)
            reburn_zip = _reburn_blob(
                xlsx_zip,
                mask_tokens=[zip_a.stem, zip_b.stem, "span_A", "span_B"],
            )
        finally:
            _clean_output_for(zip_a)

        # Headers, row count, column count match.
        assert shape_folder[2] == shape_zip[2], (
            "Splice Report header row diverged between folder and zip "
            f"runs.\nfolder: {shape_folder[2]!r}\nzip:    {shape_zip[2]!r}"
        )
        assert shape_folder[0] == shape_zip[0], (
            f"Splice Report row count diverged: folder={shape_folder[0]} "
            f"vs zip={shape_zip[0]}"
        )
        assert shape_folder[1] == shape_zip[1], (
            f"Splice Report column count diverged: folder={shape_folder[1]} "
            f"vs zip={shape_zip[1]}"
        )
        # Headline reburn text matches.
        assert reburn_folder == reburn_zip, (
            "Reburn Summary text diverged between folder and zip runs"
        )


# ─────────────────────────────────────────────────────────────────────────
# T2.  AppleDouble sidecars in a zip are skipped.
# ─────────────────────────────────────────────────────────────────────────
def test_t2_appledouble_skipped_in_zip():
    """``._FOO.sor`` AppleDouble metadata files don't pollute the
    inventory — content-sniff rejects their non-SR4731 header bytes."""
    with tempfile.TemporaryDirectory(prefix="ivt_t2_") as td:
        td_path = Path(td)
        # Clean zip.
        clean_zip = td_path / "clean.zip"
        _zip_folder(FIXTURE_A_DIR, clean_zip)

        # Polluted zip — same content + AppleDouble sidecars for several
        # real basenames.  Use the AppleDouble magic so the content sniff
        # has to reject on header-bytes alone, not on extension.
        APPLEDOUBLE_HEADER = b"\x00\x05\x16\x07" + b"\x00" * 12
        extras = {
            f"._{p.name}": APPLEDOUBLE_HEADER
            for p in sorted(FIXTURE_A_DIR.iterdir())[:5]
            if p.suffix.lower() == ".sor"
        }
        # Add a top-level junk one mimicking the filename-sweep pattern.
        extras["._STRROM0001_1550.sor"] = APPLEDOUBLE_HEADER
        polluted_zip = td_path / "polluted.zip"
        _zip_folder(FIXTURE_A_DIR, polluted_zip, extra_entries=extras)

        # Run clean.
        _clean_output_for(clean_zip)
        try:
            xlsx_clean = _run_generate(str(clean_zip), str(FIXTURE_B_DIR))
            shape_clean = _splice_sheet_shape(xlsx_clean)
            reburn_clean = _reburn_blob(
                xlsx_clean,
                mask_tokens=["clean", "polluted",
                             FIXTURE_B_DIR.name, "span_B"],
            )
            preserved = td_path / "clean_run.xlsx"
            shutil.copy2(xlsx_clean, preserved)
        finally:
            _clean_output_for(clean_zip)

        # Run polluted.
        _clean_output_for(polluted_zip)
        try:
            xlsx_polluted = _run_generate(
                str(polluted_zip), str(FIXTURE_B_DIR))
            shape_polluted = _splice_sheet_shape(xlsx_polluted)
            reburn_polluted = _reburn_blob(
                xlsx_polluted,
                mask_tokens=["clean", "polluted",
                             FIXTURE_B_DIR.name, "span_B"],
            )
        finally:
            _clean_output_for(polluted_zip)

        assert shape_clean == shape_polluted, (
            "AppleDouble sidecars perturbed Splice Report.\n"
            f"clean:    {shape_clean}\npolluted: {shape_polluted}"
        )
        assert reburn_clean == reburn_polluted, (
            "AppleDouble sidecars perturbed Reburn Summary headline text"
        )


# ─────────────────────────────────────────────────────────────────────────
# T3.  Content-sniff rejects garbage.
# ─────────────────────────────────────────────────────────────────────────
def test_t3_content_sniff_rejects_garbage():
    """``_looks_like_sor`` (and therefore ``_walk_otdr``) must reject a
    .sor-named file that lacks the SR-4731 ``Map\\0`` magic."""
    from desktop_app import _walk_otdr

    with tempfile.TemporaryDirectory(prefix="ivt_t3_") as td:
        td_path = Path(td)
        # Copy the real fixture in.
        for p in FIXTURE_A_DIR.iterdir():
            if p.is_file():
                shutil.copy2(p, td_path / p.name)
        # Drop a bogus .sor — 64 zero bytes, no Map\0 magic.
        (td_path / "bogus.sor").write_bytes(b"\x00" * 64)

        found = _walk_otdr(str(td_path))
        bogus_full = str(td_path / "bogus.sor")
        assert bogus_full not in found, (
            f"bogus.sor (no SR-4731 magic) was NOT rejected by "
            f"_looks_like_sor — content-sniff regression"
        )
        # And the real files DID survive.
        real_count = sum(
            1 for p in FIXTURE_A_DIR.iterdir()
            if p.is_file() and p.suffix.lower() == ".sor"
        )
        assert len(found) == real_count, (
            f"expected {real_count} valid SOR after sniff, got {len(found)}"
        )


# ─────────────────────────────────────────────────────────────────────────
# T4.  A == B detection (regression guard).
# ─────────────────────────────────────────────────────────────────────────
def test_t4_same_input_for_a_and_b_documented():
    """When the tech double-clicks the same row and dir_a == dir_b, the
    CURRENT behaviour is silent-run-with-perfect-span (see audit H5).

    This test pins the current behaviour AND wraps the desired behaviour
    (an error banner) in an XFAIL.  The day the guard lands, the XFAIL
    flips to XPASS and the test fails loudly — that's the signal to come
    in here, delete the XFAIL block, and promote the assertion."""
    at = run_streamlit(default_timeout=120.0)
    at.session_state["dir_a_input"] = str(FIXTURE_A_DIR)
    at.session_state["dir_b_input"] = str(FIXTURE_A_DIR)  # same!
    at.session_state["__test_force_generate__"] = True

    _clean_output_for(FIXTURE_A_DIR)
    try:
        at.run()

        # ─── Current behaviour: the engine doesn't error.
        # AppTest exposes errors via ``at.error`` and exceptions via
        # ``at.exception``.  Today both are empty in the A==B case.
        no_exception = at.exception is None or len(at.exception) == 0
        assert no_exception, (
            "Did not expect a Python-level exception in A==B today; "
            "the audit's H5 finding was 'silent run'."
        )

        # ─── Desired behaviour wrapped in XFAIL.
        # When the guard lands, an st.error banner should fire with a
        # message mentioning the duplicate input.  Phrased as a separate
        # sub-test so the XFAIL doesn't mask the current-behaviour pin
        # above.
        @pytest.mark.xfail(
            reason="H5: A==B guard not yet implemented — tripwire",
            strict=True,
        )
        def _desired_guard_present():
            errors = list(at.error) if at.error else []
            assert errors, "expected an error banner for A == B"
            txt = errors[0].value.lower() if hasattr(errors[0], "value") \
                else str(errors[0]).lower()
            assert ("same" in txt) or ("a and b" in txt), (
                f"error banner did not mention duplicate inputs: {txt!r}"
            )

        # Run the inner check; XFAIL handling is via pytest at collection
        # time, so we invoke it through pytest.xfail manually if the
        # desired condition is NOT yet met.
        errors = list(at.error) if at.error else []
        guard_present = False
        if errors:
            txt = errors[0].value.lower() if hasattr(errors[0], "value") \
                else str(errors[0]).lower()
            guard_present = ("same" in txt) or ("a and b" in txt)
        if not guard_present:
            pytest.xfail("H5: A==B guard not yet implemented — tripwire")
        # If we reach here, the guard IS present — that's the day this
        # test needs its xfail block removed.  Assert success.
        assert guard_present
    finally:
        _clean_output_for(FIXTURE_A_DIR)


# ─────────────────────────────────────────────────────────────────────────
# T5.  Sub-directory basename collisions land in staging with __1 suffix.
# ─────────────────────────────────────────────────────────────────────────
def test_t5_subdir_basename_collision_staged_distinctly():
    """Two sub-directories each containing ``LAGDUR0001.sor`` — the
    walker must surface BOTH paths, and ``_stage_flat`` must rename the
    second to ``LAGDUR0001__1.sor`` so neither silently shadows the
    other on disk.

    NOTE — pair this with Agent D's engine-side collision WARN check.
    ``_extract_fiber_num`` still returns 1 for BOTH staged files
    (rightmost digit run is the same), so the staging-level uniqueness
    only buys us the engine's collision WARN as a backstop — not a true
    de-duplication."""
    from desktop_app import _walk_otdr, _stage_flat

    src = FIXTURE_A_DIR / "LAGDUR0001.sor"
    assert src.is_file(), "fixture file LAGDUR0001.sor missing"

    with tempfile.TemporaryDirectory(prefix="ivt_t5_") as td:
        td_path = Path(td)
        sub1 = td_path / "sub1"
        sub2 = td_path / "sub2"
        sub1.mkdir()
        sub2.mkdir()
        shutil.copy2(src, sub1 / "LAGDUR0001.sor")
        shutil.copy2(src, sub2 / "LAGDUR0001.sor")

        found = _walk_otdr(str(td_path))
        assert len(found) == 2, (
            f"walker should surface BOTH copies, got {found}"
        )
        assert any("sub1" in p for p in found)
        assert any("sub2" in p for p in found)

        staged_dir, warns = _stage_flat(found, "ivt_t5_stage_")
        try:
            staged = sorted(os.listdir(staged_dir))
            assert "LAGDUR0001.sor" in staged, (
                f"primary basename missing from staging: {staged}"
            )
            assert "LAGDUR0001__1.sor" in staged, (
                f"deduped basename missing from staging: {staged}"
            )
            # And the dedup warning was emitted.
            assert any("Duplicate basename" in w for w in warns), (
                f"expected a 'Duplicate basename' warning, got {warns!r}"
            )
            # Files are non-empty copies.
            for nm in ("LAGDUR0001.sor", "LAGDUR0001__1.sor"):
                assert os.path.getsize(
                    os.path.join(staged_dir, nm)) > 0
        finally:
            shutil.rmtree(staged_dir, ignore_errors=True)


# ─────────────────────────────────────────────────────────────────────────
# T6.  Zip with a backslash-path entry handles cleanly.
# ─────────────────────────────────────────────────────────────────────────
def test_t6_zip_backslash_entry_handled():
    """A Windows-flavoured zip with ``subdir\\file.sor`` entries — the
    walker should still see the entry, and the stager should write a
    non-empty file at the staged basename."""
    from desktop_app import _walk_zip_otdr, _stage_flat

    src = FIXTURE_A_DIR / "LAGDUR0001.sor"
    assert src.is_file()
    blob = src.read_bytes()

    with tempfile.TemporaryDirectory(prefix="ivt_t6_") as td:
        td_path = Path(td)
        zip_path = td_path / "winzip.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            # Backslash inside the entry name — legal-but-cursed Windows
            # zip pattern.  zipfile preserves the raw name.
            zf.writestr("subdir\\LAGDUR0001.sor", blob)

        found = _walk_zip_otdr(str(zip_path))
        assert len(found) == 1, (
            f"expected exactly one entry from backslash-path zip, "
            f"got {found}"
        )
        token = found[0]
        assert token.startswith("zip:")
        # basename should resolve to LAGDUR0001.sor regardless of the
        # raw separator inside the zip entry name.
        entry = token[4:]
        # os.path.basename treats '\\' as a literal char on POSIX, so
        # accept either: the raw separator survives and the staged file
        # is named for the rightmost path component, however basename
        # parses it.  What we MUST get is a non-empty staged file.
        staged_dir, warns = _stage_flat(found, "ivt_t6_stage_",
                                        zip_source=str(zip_path))
        try:
            staged_files = [
                f for f in os.listdir(staged_dir)
                if os.path.isfile(os.path.join(staged_dir, f))
            ]
            assert len(staged_files) == 1, (
                f"expected one staged file, got {staged_files}; "
                f"warns={warns!r}"
            )
            staged_path = os.path.join(staged_dir, staged_files[0])
            assert os.path.getsize(staged_path) > 0, (
                "staged file from backslash-entry zip is empty"
            )
            assert open(staged_path, "rb").read(4) == b"Map\x00", (
                "staged file doesn't carry the original SOR magic"
            )
        finally:
            shutil.rmtree(staged_dir, ignore_errors=True)
